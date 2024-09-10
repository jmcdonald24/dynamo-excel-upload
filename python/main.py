import boto3
import openpyxl
from io import BytesIO

session = boto3.Session(profile_name='default')
ddb_client = session.client('dynamodb')
s3_client = session.client('s3')

item = {
    'id': {'N': '1'},
    'name': {'S': 'John Doe'}
}

def write_to_dynamodb(client, table_name, item: dict):
    response = ddb_client.put_item(
        TableName=table_name,
        Item=item
    )
    
    print(response)


def delete_all_rows(table_name):


    TABLE = table_name
    ID = 'id'

    table = boto3.resource('dynamodb').Table(TABLE)

    scan = None

    with table.batch_writer() as batch:
        count = 0
        while scan is None or 'LastEvaluatedKey' in scan:
            if scan is not None and 'LastEvaluatedKey' in scan:
                scan = table.scan(
                    ProjectionExpression=ID,
                    ExclusiveStartKey=scan['LastEvaluatedKey'],
                )
            else:
                scan = table.scan(ProjectionExpression=ID)

            for item in scan['Items']:
                if count % 5000 == 0:
                    print(count)
                batch.delete_item(Key={ID: item[ID]})
                count = count + 1

def compare_header_with_fixed_values(file_path, fixed_values):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    header_row = sheet[1]
    header_values = [cell.value for cell in header_row]

    if header_values == fixed_values:
        print("Header matches fixed values")
    else:
        print("Header does not match fixed values")

def read_excel_from_s3(bucket_name, s3_file_key):
    s3 = boto3.client('s3')
    excel_data = s3.get_object(Bucket=bucket_name, Key=s3_file_key)['Body'].read()
    workbook = openpyxl.load_workbook(BytesIO(excel_data), data_only=True)
    sheet = workbook.active

    return sheet

def write_sheet_to_dynamodb(sheet, table_name):
    header_row = sheet[1]
    header_values = [cell.value for cell in header_row]

    for row in sheet.iter_rows(min_row=2):

        item = {}
        for index, cell in enumerate(row):
            field_name = header_values[index]
            field_value = cell.value



            item[field_name] = {'S': str(field_value)}
        print(item)
        write_to_dynamodb(ddb_client, table_name, item)
        
        
def read_excel_and_write_to_dynamodb(file_path, table_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    header_row = sheet[1]
    header_values = [cell.value for cell in header_row]



    for row in sheet.iter_rows(min_row=2):




        item = {}
        for index, cell in enumerate(row):
            field_name = header_values[index]
            field_value = cell.value



            item[field_name] = {'S': str(field_value)}
        print(item)
        write_to_dynamodb(ddb_client, table_name, item)


# Example usage:
#fixed_values = ["Name", "Email", "Age"]
#compare_header_with_fixed_values("Book1.xlsx", fixed_values)


# for i in range(10):
#     write_to_dynamodb(ddb_client, 'example_table', generate_random_ddbdata())

#delete_all_rows('example_table')

def download_excel_from_s3(bucket_name, s3_file_key, local_file_path):
    s3 = boto3.client('s3')
    s3.download_file(bucket_name, s3_file_key, local_file_path)
    print(f"Downloaded {s3_file_key} from bucket {bucket_name} to {local_file_path}")

def read_excel_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1):
        row_values = [cell.value for cell in row]
        


bucket_name = 'dynamorawdatabucket'
s3_file_key = 'Book1.xlsx'


#download_excel_from_s3(bucket_name, s3_file_key, local_file_path)
#read_excel_file(local_file_path)


#read_excel_and_write_to_dynamodb("Book1.xlsx", "example_table")

data_sheet = read_excel_from_s3(bucket_name, s3_file_key)
write_sheet_to_dynamodb(data_sheet, "example_table")